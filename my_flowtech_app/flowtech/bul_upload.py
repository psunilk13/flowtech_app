import frappe
import csv
import os
import openpyxl
from frappe.utils.file_manager import get_files_path, get_file_path


# =====================================================
# BULK UPLOAD ITEMS (CSV / EXCEL)
# =====================================================
@frappe.whitelist()
def upload_bulk_items(parent, file_url):

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = os.path.join(get_files_path(), os.path.basename(file_doc.file_name))

    if not os.path.exists(file_path):
        frappe.throw(f"File not found: {file_path}")

    parent_doc = frappe.get_doc("Enquiry", parent)
    count = 0

    def get_warehouse(item_code):
        bin_doc = frappe.get_all(
            "Bin",
            filters={"item_code": item_code},
            fields=["warehouse", "actual_qty"]
        )
        if bin_doc:
            return bin_doc[0]["warehouse"], bin_doc[0]["actual_qty"]
        return "", 0

    ext = os.path.splitext(file_path)[1].lower()

    # ---------------- CSV ----------------
    if ext == ".csv":
        with open(file_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if not row.get("item"):
                    continue

                warehouse, warehouse_qty = get_warehouse(row["item"])

                parent_doc.append("items_details", {
                    "item": row.get("item"),
                    "item_name": row.get("item_name"),
                    "quantity": row.get("quantity"),
                    "actual_price": row.get("actual_price"),
                    "warehouse": warehouse,
                    "warehouse_qty": warehouse_qty
                })
                count += 1

    # ---------------- EXCEL ----------------
    elif ext == ".xlsx":
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]

        required = {"item", "item_name", "quantity", "actual_price"}
        if not required.issubset(headers):
            frappe.throw("Excel missing required columns")

        idx = {h: headers.index(h) for h in headers}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[idx["item"]]:
                continue

            warehouse, warehouse_qty = get_warehouse(row[idx["item"]])

            parent_doc.append("items_details", {
                "item": row[idx["item"]],
                "item_name": row[idx["item_name"]],
                "quantity": row[idx["quantity"]],
                "actual_price": row[idx["actual_price"]],
                "warehouse": warehouse,
                "warehouse_qty": warehouse_qty
            })
            count += 1

    else:
        frappe.throw("Upload only CSV or XLSX")

    parent_doc.save(ignore_permissions=True)
    frappe.db.commit()

    return f"Uploaded {count} items successfully"

# # new bulkupload working code
# # -----
# import frappe
# import csv
# import os
# import openpyxl
# from frappe.utils.file_manager import get_files_path


# @frappe.whitelist()
# def upload_bulk_items(parent, file_url):
#     """
#     Bulk upload items into Enquiry -> items_details child table
#     GST is a Link to GST Rates DocType
#     """

#     # ---------------- FILE ---------------- #
#     file_doc = frappe.get_doc("File", {"file_url": file_url})
#     file_path = os.path.join(get_files_path(), file_doc.file_name)

#     if not os.path.exists(file_path):
#         frappe.throw(f"File not found: {file_path}")

#     parent_doc = frappe.get_doc("Enquiry", parent)
#     count = 0

#     # ---------------- HELPERS ---------------- #

#     def normalize_gst(gst):
#         """
#         Normalize GST to valid GST Rates DocType name
#         """
#         if gst in (None, "", 0):
#             gst_value = "0%"
#         else:
#             try:
#                 gst = float(gst)
#                 if gst < 1:
#                     gst = int(round(gst * 100))
#                 else:
#                     gst = int(round(gst))
#                 gst_value = f"{gst}%"
#             except Exception:
#                 gst_value = str(gst).strip()

#         if not frappe.db.exists("GST Rates", gst_value):
#             frappe.throw(f"Invalid GST Rate in file: {gst_value}")

#         return gst_value

#     def get_bin_qty(item_code, warehouse):
#         return frappe.db.get_value(
#             "Bin",
#             {"item_code": item_code, "warehouse": warehouse},
#             "actual_qty"
#         ) or 0

#     def get_warehouse_from_bin(item_code):
#         bins = frappe.get_all(
#             "Bin",
#             filters={"item_code": item_code, "actual_qty": (">", 0)},
#             fields=["warehouse", "actual_qty"],
#             order_by="actual_qty desc"
#         )
#         if bins:
#             return bins[0]["warehouse"], bins[0]["actual_qty"]
#         return "", 0

#     # ---------------- PROCESS FILE ---------------- #

#     file_ext = os.path.splitext(file_path)[1].lower()

#     # ---------- CSV ---------- #
#     if file_ext == ".csv":
#         with open(file_path, "r", encoding="utf-8") as f:
#             reader = csv.DictReader(f)

#             for row in reader:
#                 item_code = row.get("item")
#                 if not item_code:
#                     continue

#                 warehouse = row.get("warehouse")
#                 if warehouse:
#                     warehouse_qty = get_bin_qty(item_code, warehouse)
#                 else:
#                     warehouse, warehouse_qty = get_warehouse_from_bin(item_code)

#                 parent_doc.append("items_details", {
#                     "custom_serial_no": row.get("custom_serial_no"),
#                     "item": item_code,
#                     "item_name": row.get("item_name"),
#                     "quantity": row.get("quantity"),
#                     "actual_price": row.get("actual_price"),
#                     # "discount": row.get("discount"),
#                     "gst": normalize_gst(row.get("gst")),
#                     "warehouse": warehouse,
#                     "warehouse_qty": warehouse_qty
#                 })
#                 count += 1

#     # ---------- EXCEL ---------- #
#     elif file_ext == ".xlsx":
#         wb = openpyxl.load_workbook(file_path)
#         sheet = wb.active

#         # headers = [str(cell.value).strip() for cell in sheet[1]]
# 		headers = [
#     str(cell.value).strip().lower() if cell.value else ""
#     for cell in sheet[1]
# ]

#         # idx = {h: i for i, h in enumerate(headers)}
# 		idx = {h.lower(): i for i, h in enumerate(headers) if h}


#         # required_cols = {
#         #     "custom_serial_no", "item", "item_name",
#         #     "quantity", "actual_price", "discount", "gst"
#         # }

# 		required_cols = {
#     "custom_serial_no", "item", "item_name",
#     "quantity", "actual_price", "gst"
# }

#         missing = required_cols - set(idx)
#         if missing:
#             frappe.throw(f"Missing columns: {', '.join(missing)}")

#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             if not any(row):
#                 continue

#             item_code = row[idx["item"]]
#             if not item_code:
#                 continue

#             excel_warehouse = row[idx["warehouse"]] if "warehouse" in idx else None

#             if excel_warehouse:
#                 warehouse = excel_warehouse
#                 warehouse_qty = get_bin_qty(item_code, warehouse)
#             else:
#                 warehouse, warehouse_qty = get_warehouse_from_bin(item_code)

#             parent_doc.append("items_details", {
#                 "custom_serial_no": row[idx["custom_serial_no"]],
#                 "item": item_code,
#                 "item_name": row[idx["item_name"]],
#                 "quantity": row[idx["quantity"]],
#                 "actual_price": row[idx["actual_price"]],
#                 # "discount": row[idx["discount"]],
#                 "gst": normalize_gst(row[idx["gst"]]),
#                 "warehouse": warehouse,
#                 "warehouse_qty": warehouse_qty
#             })
#             count += 1

#     else:
#         frappe.throw("Only CSV and XLSX files are supported")

#     parent_doc.save(ignore_permissions=True)
#     frappe.db.commit()

#     return f"✅ Successfully uploaded {count} items"


import frappe
from erpnext.stock.doctype.delivery_note.delivery_note import DeliveryNote

class CustomDeliveryNote(DeliveryNote):
    def autoname(self):
        from frappe.utils import nowdate

        date = nowdate()
        year = int(date[0:4])
        month = int(date[5:7])

        if month >= 4:
            start_year = year
            end_year = str(year + 1)[-2:]
        else:
            start_year = year - 1
            end_year = str(year)[-2:]

        fy = f"{start_year}-{end_year}"

        last = frappe.db.sql("""
            SELECT name FROM `tabDelivery Note`
            WHERE name LIKE %s
            ORDER BY name DESC
            LIMIT 1
        """, (f"DC{fy}/%",))

        if last:
            last_no = int(last[0][0].split("/")[-1])
            new_no = str(last_no + 1).zfill(5)
        else:
            new_no = "00001"

        self.name = f"DC{fy}/{new_no}"

#---------------Managing Item Cancelling code-----------------
import frappe

@frappe.whitelist()
def cancel_oa_items(oa_name, items):

    import json
    items = json.loads(items)

    # =========================
    # 🔹 GET DOCUMENTS
    # =========================
    oa = frappe.get_doc("OA Register", oa_name)

    ri_name = frappe.db.get_value("Reserved Inventory", {"oa_register": oa_name})
    ri = frappe.get_doc("Reserved Inventory", ri_name) if ri_name else None

    # FLAGS
    oa.flags.ignore_validate_update_after_submit = True
    oa.flags.ignore_validate = True

    if ri:
        ri.flags.ignore_validate_update_after_submit = True
        ri.flags.ignore_validate = True

    oa_rows = oa.get("items_details") or []
    ri_rows = ri.get("items") if ri else []

    # =========================
    # 🔹 PROCESS CANCEL
    # =========================
    for item in items:

        serial = str(item.get("serial_no")).strip()
        cancel_qty = float(item.get("cancel_qty") or 0)

        parent_old_qty = 0

        # 🔹 GET PARENT QTY
        for row in oa_rows:
            if str(row.custom_serial_no).strip() == serial:
                parent_old_qty = row.quantity or 0
                break

        # =========================
        # 🔹 OA UPDATE
        # =========================
        for row in oa_rows:

            row_serial = str(row.custom_serial_no).strip()

            # 🔹 PARENT
            if row_serial == serial:

                if cancel_qty > parent_old_qty:
                    frappe.throw(f"Cancel qty exceeds available qty for {serial}")

                new_qty = parent_old_qty - cancel_qty
                row.set("quantity", new_qty)

                # 🔍 CHECK CHILD
                has_child = any(
                    str(r.custom_serial_no).strip().startswith(serial + "-")
                    for r in oa_rows
                )

                if has_child:
                    row.set("reserved_qty", 0)
                    row.already_reserved = 0
                    row.db_set("already_reserved", 0)
                else:
                    row.set("reserved_qty", new_qty)
                    row.already_reserved = new_qty
                    row.db_set("already_reserved", new_qty)

            # 🔹 CHILD
            elif row_serial.startswith(serial + "-"):

                old_child_qty = row.quantity or 0
                old_reserved = row.reserved_qty or 0
                old_remaining = row.remaining_quantity or 0

                if parent_old_qty > 0:
                    reduce_qty = (old_child_qty / parent_old_qty) * cancel_qty
                else:
                    reduce_qty = 0

                new_child_qty = old_child_qty - reduce_qty
                new_reserved = max(old_reserved - reduce_qty, 0)
                new_remaining = old_remaining + reduce_qty

                row.set("quantity", new_child_qty)
                row.set("reserved_qty", new_reserved)

                # 🔥 IMPORTANT FIELD
                row.already_reserved = new_reserved
                row.db_set("already_reserved", new_reserved)

                row.set("remaining_quantity", new_remaining)

        # =========================
        # 🔹 RESERVED INVENTORY UPDATE
        # =========================
        if ri:

            for row in ri_rows:

                row_serial = str(row.custom_serial_no).strip()

                if row_serial == serial:

                    new_qty = (row.quantity or 0) - cancel_qty
                    row.set("quantity", new_qty)

                    has_child = any(
                        str(r.custom_serial_no).strip().startswith(serial + "-")
                        for r in ri_rows
                    )

                    if has_child:
                        row.set("reserved_qty", 0)
                    else:
                        row.set("reserved_qty", new_qty)

                elif row_serial.startswith(serial + "-"):

                    old_child_qty = row.quantity or 0

                    if parent_old_qty > 0:
                        reduce_qty = (old_child_qty / parent_old_qty) * cancel_qty
                    else:
                        reduce_qty = 0

                    new_child_qty = old_child_qty - reduce_qty
                    new_reserved = (row.reserved_qty or 0) - reduce_qty

                    row.set("quantity", new_child_qty)
                    row.set("reserved_qty", new_reserved)

    # =========================
    # 🔥 DELETE ZERO QTY ROWS
    # =========================
    oa.set("items_details", [d for d in oa.get("items_details") if (d.quantity or 0) > 0])

    if ri:
        ri.set("items", [d for d in ri.get("items") if (d.quantity or 0) > 0])

    # =========================
    # 🔥 DELETE DOCUMENT IF EMPTY
    # =========================
    if not oa.get("items_details"):

        # 🔥 CANCEL + DELETE OA
        if oa.docstatus == 1:
            oa.cancel()

        frappe.delete_doc("OA Register", oa.name, force=1)

        # 🔥 CANCEL + DELETE RI
        if ri:
            if ri.docstatus == 1:
                ri.cancel()

            frappe.delete_doc("Reserved Inventory", ri.name, force=1)

        return "All items cancelled. Documents deleted"
        
    # =========================
    # 🔹 SAVE
    # =========================
    oa.save()

    if ri:
        ri.save()

    return "Items Cancelled Successfully"\

